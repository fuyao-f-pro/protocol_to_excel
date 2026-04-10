"""
excel_writer.py — 将解析结果写入标准 Excel（完整版）
"""
import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from models import ParsedDocument, DataPoint


class ExcelWriter:

    def __init__(self, doc: ParsedDocument):
        self.doc = doc
        self.wb = Workbook()

    def write(self, output_path: str) -> str:
        ws = self.wb.active
        ws.title = "点位表"

        self._write_info_row(ws)

        header_row = 3
        self._write_header(ws, header_row)

        data_start = header_row + 1
        self._write_data(ws, data_start)

        total_rows = data_start + len(self.doc.points) - 1
        if len(self.doc.points) == 0:
            total_rows = data_start
        self._apply_styles(ws, header_row, data_start, total_rows)
        self._auto_width(ws)
        ws.freeze_panes = config.STYLE.get("freeze", "C4")

        self._write_summary_sheet()

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        self.wb.save(output_path)
        return output_path

    # ----------------------------------------------------------
    def _write_info_row(self, ws):
        info = (
            f"文件: {self.doc.filename}  |  "
            f"协议: {self.doc.protocol_type}  |  "
            f"设备: {self.doc.device_name}  |  "
            f"点位数: {len(self.doc.points)}"
        )
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(config.STANDARD_COLUMNS),
        )
        cell = ws.cell(row=1, column=1, value=info)
        cell.font = Font(bold=True, size=12, color="2F5496")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30

    # ----------------------------------------------------------
    def _write_header(self, ws, row: int):
        header_fill = PatternFill(
            start_color=config.STYLE["header_bg"],
            end_color=config.STYLE["header_bg"],
            fill_type="solid",
        )
        header_font = Font(
            bold=True,
            color=config.STYLE["header_font_color"],
            size=config.STYLE["header_font_size"],
        )
        header_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
        thin_border = self._border()

        for col_idx, col_name in enumerate(config.STANDARD_COLUMNS, 1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        ws.row_dimensions[row].height = 28
        # 添加筛选
        last_col = get_column_letter(len(config.STANDARD_COLUMNS))
        ws.auto_filter.ref = f"A{row}:{last_col}{row}"

    # ----------------------------------------------------------
    def _write_data(self, ws, start_row: int):
        data_font = Font(size=config.STYLE["data_font_size"])
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = self._border()

        # 需要居中的列索引（0-based）
        center_cols = {0, 3, 5, 6, 7, 8, 9, 10, 12, 13, 15, 16}  # 序号/类型/地址等

        colors = config.STYLE["alt_row_colors"]

        for r_idx, point in enumerate(self.doc.points):
            row_num = start_row + r_idx
            row_data = point.to_row()
            bg_color = colors[r_idx % len(colors)]
            row_fill = PatternFill(
                start_color=bg_color, end_color=bg_color, fill_type="solid",
            )

            for c_idx, value in enumerate(row_data):
                cell = ws.cell(row=row_num, column=c_idx + 1, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.fill = row_fill
                cell.alignment = center_align if c_idx in center_cols else left_align

            ws.row_dimensions[row_num].height = 22

    # ----------------------------------------------------------
    def _apply_styles(self, ws, header_row, data_start, total_rows):
        """对特定列做条件格式高亮等"""
        # 开关量行标红
        type_col = config.STANDARD_COLUMNS.index("点位类型") + 1
        red_font = Font(color="C00000", bold=True, size=config.STYLE["data_font_size"])
        for r in range(data_start, total_rows + 1):
            cell = ws.cell(row=r, column=type_col)
            if cell.value == "开关量":
                cell.font = red_font

    # ----------------------------------------------------------
    def _auto_width(self, ws):
        """自动调整列宽"""
        min_width = 8
        max_width = 40
        for col_idx in range(1, len(config.STANDARD_COLUMNS) + 1):
            col_letter = get_column_letter(col_idx)
            best = min_width
            for row in ws.iter_rows(min_row=3, max_col=col_idx, min_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        length = len(str(cell.value))
                        # 中文字符按1.8倍宽度
                        cn_count = sum(
                            1 for ch in str(cell.value)
                            if '\u4e00' <= ch <= '\u9fff'
                        )
                        width = length + cn_count * 0.8 + 2
                        if width > best:
                            best = width
            ws.column_dimensions[col_letter].width = min(best, max_width)

    # ----------------------------------------------------------
    def _write_summary_sheet(self):
        """汇总统计 Sheet"""
        ws2 = self.wb.create_sheet(title="汇总统计")
        points = self.doc.points

        # 统计
        total = len(points)
        analog_count = sum(1 for p in points if p.点位类型 == "模拟量")
        digital_count = sum(1 for p in points if p.点位类型 == "开关量")
        r_count = sum(1 for p in points if p.访问规则 == "R")
        rw_count = sum(1 for p in points if p.访问规则 == "RW")
        w_count = sum(1 for p in points if p.访问规则 == "W")

        # 数据类型统计
        dtype_stats = {}
        for p in points:
            dtype_stats[p.数据类型] = dtype_stats.get(p.数据类型, 0) + 1

        # 写入
        header_fill = PatternFill(
            start_color=config.STYLE["header_bg"],
            end_color=config.STYLE["header_bg"],
            fill_type="solid",
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        bold_font = Font(bold=True, size=11)
        normal_font = Font(size=11)
        thin = self._border()
        center = Alignment(horizontal="center", vertical="center")

        # 基本信息
        summary_data = [
            ["项目", "值"],
            ["文件名", self.doc.filename],
            ["协议类型", self.doc.protocol_type],
            ["设备名称", self.doc.device_name],
            ["总点位数", total],
            ["模拟量", analog_count],
            ["开关量", digital_count],
            ["只读(R)", r_count],
            ["读写(RW)", rw_count],
            ["只写(W)", w_count],
        ]

        for r_idx, row_data in enumerate(summary_data):
            for c_idx, val in enumerate(row_data):
                cell = ws2.cell(row=r_idx + 1, column=c_idx + 1, value=val)
                cell.border = thin
                cell.alignment = center
                if r_idx == 0:
                    cell.fill = header_fill
                    cell.font = header_font
                elif c_idx == 0:
                    cell.font = bold_font
                else:
                    cell.font = normal_font

        # 数据类型统计
        start_row = len(summary_data) + 2
        ws2.cell(row=start_row, column=1, value="数据类型").fill = header_fill
        ws2.cell(row=start_row, column=1).font = header_font
        ws2.cell(row=start_row, column=1).border = thin
        ws2.cell(row=start_row, column=2, value="数量").fill = header_fill
        ws2.cell(row=start_row, column=2).font = header_font
        ws2.cell(row=start_row, column=2).border = thin

        for i, (dtype, cnt) in enumerate(sorted(dtype_stats.items()), 1):
            ws2.cell(row=start_row + i, column=1, value=dtype).border = thin
            ws2.cell(row=start_row + i, column=1).font = normal_font
            ws2.cell(row=start_row + i, column=2, value=cnt).border = thin
            ws2.cell(row=start_row + i, column=2).font = normal_font

        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 30

    # ----------------------------------------------------------
    @staticmethod
    def _border():
        side = Side(style="thin", color="B4C6E7")
        return Border(left=side, right=side, top=side, bottom=side)